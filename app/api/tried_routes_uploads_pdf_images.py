from fastapi import APIRouter, UploadFile, File, HTTPException, BackgroundTasks, Header, Path, Depends, status, Form
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from typing import Optional, Any, Dict, List
from pathlib import Path as FilePath
from pydantic import BaseModel, Field, validator
from app.models.UserPlan import UserPlan
from app.models.UsageLog import UsageLog

import shutil
import json
import io
import csv
import tempfile
import zipfile
import base64
import uuid
import datetime
import mimetypes
import os

import fitz

from openpyxl import Workbook

from sqlalchemy.orm import Session
from sqlalchemy import or_
from app.core.database import get_db
from app.models.project import Project
from app.models.project_file import ProjectFile

router = APIRouter(prefix="/uploads-pdf-images", tags=["PDF → Images"])

BASE_DIR = FilePath(__file__).resolve().parents[2] if "__file__" in globals() else FilePath(".")
UPLOAD_ROOT = BASE_DIR / "uploads"
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

RENDER_SCALE_PAGE = 3.0
RENDER_SCALE_REGION = 3.0
MERGE_IOU = 0.60
REGION_PAD_FRAC = 0.02
MAX_PIXELS_GUARD = 20000 * 20000
MIN_AREA_PCT = 0.002
MIN_SIDE_PX = 40
MAX_RECTS_PER_PAGE = 24

def ensure_dir(p: FilePath):
    p.mkdir(parents=True, exist_ok=True)

def rect_area(r: fitz.Rect):
    return max(0.0, (r.x1 - r.x0)) * max(0.0, (r.y1 - r.y0))

def iou(a, b):
    inter = fitz.Rect(max(a.x0, b.x0), max(a.y0, b.y0), min(a.x1, b.x1), min(a.y1, b.y1))
    ia = rect_area(inter)
    if ia <= 0:
        return 0.0
    ua = rect_area(a) + rect_area(b) - ia
    return ia / ua if ua > 0 else 0.0

def pad_rect(r: fitz.Rect, pad: float, page: fitz.Page):
    pr = page.rect
    p = fitz.Rect(r.x0 - pad, r.y0 - pad, r.x1 + pad, r.y1 + pad)
    p.x0 = max(p.x0, pr.x0)
    p.y0 = max(p.y0, pr.y0)
    p.x1 = min(p.x1, pr.x1)
    p.y1 = min(p.y1, pr.y1)
    return p

def merge_near_duplicates(rects, thr):
    rects = sorted(rects, key=lambda r: rect_area(r), reverse=True)
    kept = []
    for r in rects:
        if any(iou(r, k) >= thr for k in kept):
            continue
        kept.append(r)
    return kept

def get_table_rects(page: fitz.Page) -> List[fitz.Rect]:
    table_rects: List[fitz.Rect] = []

    try:
        tables = page.find_tables()
        for table in tables:
            try:
                r = fitz.Rect(table.bbox)

                if r.is_empty or r.is_infinite:
                    continue

                if r.width < 60 or r.height < 35:
                    continue

                table_rects.append(r)
            except Exception:
                pass
    except Exception:
        pass

    return table_rects


def get_vector_diagram_rects(page: fitz.Page) -> List[fitz.Rect]:
    diagram_rects: List[fitz.Rect] = []

    # Better than get_drawings() because it groups vector lines into diagram-like boxes
    try:
        clusters = page.cluster_drawings()
        for r in clusters:
            rr = fitz.Rect(r)

            if rr.is_empty or rr.is_infinite:
                continue

            if rr.width < 45 or rr.height < 45:
                continue

            diagram_rects.append(rr)
    except Exception:
        pass

    return diagram_rects

def get_image_block_rects(page: fitz.Page) -> List[fitz.Rect]:
    image_rects: List[fitz.Rect] = []

    try:
        data = page.get_text("dict")
        blocks = data.get("blocks", [])

        for block in blocks:
            try:
                if block.get("type") != 1:
                    continue

                bbox = block.get("bbox")
                if not bbox:
                    continue

                r = fitz.Rect(bbox)

                if r.is_empty or r.is_infinite:
                    continue

                if r.width < 50 or r.height < 50:
                    continue

                image_rects.append(r)
            except Exception:
                pass
    except Exception:
        pass

    return image_rects

def save_pixmap_png_safe(pix, out_path: FilePath):
    try:
        if getattr(pix, "w", 0) <= 0 or getattr(pix, "h", 0) <= 0:
            return False
        if pix.w * pix.h > MAX_PIXELS_GUARD:
            return False
        pix.save(str(out_path))
        return True
    except Exception:
        try:
            pix2 = fitz.Pixmap(fitz.csRGB, pix)
            pix2.save(str(out_path))
            return True
        except Exception:
            return False

def extract_images_to_manifest(pdf_path: FilePath, out_dir: FilePath, json_path: FilePath, user_id: str, base_url: str) -> Dict[str, Any]:
    ensure_dir(out_dir)
    images_dir = out_dir / "images"
    ensure_dir(images_dir)

    pdf_stem = pdf_path.stem
    doc = fitz.open(pdf_path)

    manifest: Dict[str, Any] = {
        "source_pdf": str(pdf_path),
        "source_pdf_url": f"{base_url}/uploads/{user_id}/{pdf_path.name}",
        "image_count": 0,
        "items": []
    }

    for pi, page in enumerate(doc, 1):
        text = page.get_text("text")
        has_text = bool(text.strip())

        try:
            dl = page.get_displaylist()
        except Exception:
            dl = None

        try:
            M_full = fitz.Matrix(RENDER_SCALE_PAGE, RENDER_SCALE_PAGE)
            pix_full = dl.get_pixmap(matrix=M_full, alpha=False) if dl else page.get_pixmap(matrix=M_full, alpha=False)
            full_img = images_dir / f"page_{pi}_full_page_image.png"
            save_pixmap_png_safe(pix_full, full_img)
        except Exception:
            pass

        page_area = rect_area(page.rect)
        min_side = MIN_SIDE_PX / max(RENDER_SCALE_REGION, 1e-6)
        raw: List[fitz.Rect] = []

        try:
            images = page.get_images(full=True)
        except Exception:
            images = []

        for xref, *_ in images:
            try:
                image_rects = page.get_image_rects(xref)
            except Exception:
                image_rects = []

            for r in image_rects:
                try:
                    r = fitz.Rect(r)

                    if r.is_empty or r.is_infinite:
                        continue

                    pad = max(r.width, r.height) * REGION_PAD_FRAC + 2.0
                    r2 = pad_rect(r, pad, page)

                    if rect_area(r2) < MIN_AREA_PCT * page_area:
                        continue

                    if (r2.x1 - r2.x0) < min_side or (r2.y1 - r2.y0) < min_side:
                        continue

                    if r2.width < 120 or r2.height < 120:
                        continue

                    if rect_area(r2) > page_area * 0.35:
                        continue

                    raw.append(r2)
                except Exception:
                    pass

        try:
            block_image_rects = get_image_block_rects(page)
        except Exception:
            block_image_rects = []

        for r in block_image_rects:
            try:
                r = fitz.Rect(r)

                if r.is_empty or r.is_infinite:
                    continue

                pad = max(r.width, r.height) * REGION_PAD_FRAC + 2.0
                r2 = pad_rect(r, pad, page)

                if rect_area(r2) < MIN_AREA_PCT * page_area:
                    continue

                if (r2.x1 - r2.x0) < min_side or (r2.y1 - r2.y0) < min_side:
                    continue

                if r2.width < 80 or r2.height < 80:
                    continue

                if rect_area(r2) > page_area * 0.35:
                    continue

                raw.append(r2)
            except Exception:
                pass

        try:
            table_rects = get_table_rects(page)
        except Exception:
            table_rects = []

        for r in table_rects:
            try:
                r = fitz.Rect(r)

                if r.is_empty or r.is_infinite:
                    continue

                pad = max(r.width, r.height) * REGION_PAD_FRAC + 2.0
                r2 = pad_rect(r, pad, page)

                if rect_area(r2) < MIN_AREA_PCT * page_area:
                    continue

                if (r2.x1 - r2.x0) < min_side or (r2.y1 - r2.y0) < min_side:
                    continue

                if rect_area(r2) > page_area * 0.35:
                    continue

                raw.append(r2)
            except Exception:
                pass

        try:
            diagram_rects = get_vector_diagram_rects(page)
        except Exception:
            diagram_rects = []

        for r in diagram_rects:
            try:
                r = fitz.Rect(r)

                if r.is_empty or r.is_infinite:
                    continue

                pad = max(r.width, r.height) * REGION_PAD_FRAC + 2.0
                r2 = pad_rect(r, pad, page)

                if rect_area(r2) < MIN_AREA_PCT * page_area:
                    continue

                if (r2.x1 - r2.x0) < min_side or (r2.y1 - r2.y0) < min_side:
                    continue

                if rect_area(r2) > page_area * 0.30:
                    continue

                raw.append(r2)
            except Exception:
                pass

        rects = merge_near_duplicates(raw, MERGE_IOU)
        rects = sorted(rects, key=lambda rr: (rr.y0, rr.x0))[:MAX_RECTS_PER_PAGE]

        for idx, r in enumerate(rects, 1):
            try:
                M_reg = fitz.Matrix(RENDER_SCALE_REGION, RENDER_SCALE_REGION)
                pix = dl.get_pixmap(matrix=M_reg, clip=r, alpha=False) if dl else page.get_pixmap(matrix=M_reg, clip=r, alpha=False)
                img_path = images_dir / f"p{pi}_img{idx}.png"

                if save_pixmap_png_safe(pix, img_path):
                    manifest["items"].append({
                        "page": pi,
                        "index": idx,
                        "type": "image",
                        "has_text": has_text,
                        "file_name": img_path.name,
                        "system_path": str(img_path.resolve()),
                        "web_path": f"{base_url}/uploads/{user_id}/{pdf_stem}/images/{img_path.name}",
                        "is_visible": True
                    })
                    manifest["image_count"] += 1

            except Exception:
                pass

    doc.close()

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    return {"manifest": manifest, "json_path": json_path}

async def handle_pdf_upload(
    id: str,
    file: UploadFile,
    background_tasks: Optional[BackgroundTasks],
    authorization: Optional[str],
    base_url: str,
):
    if not file.filename.lower().endswith(".pdf"):
        return {"error": "Only PDF files are allowed"}
    user_dir = UPLOAD_ROOT / id
    ensure_dir(user_dir)
    pdf_name = file.filename
    pdf_stem = FilePath(pdf_name).stem
    pdf_dir = user_dir / pdf_stem
    images_dir = pdf_dir / "images"
    dest_pdf = user_dir / pdf_name
    dest_json = user_dir / f"{pdf_stem}.json"
    if images_dir.exists():
        shutil.rmtree(images_dir, ignore_errors=True)
    ensure_dir(pdf_dir)
    with dest_pdf.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    result = extract_images_to_manifest(dest_pdf, pdf_dir, dest_json, user_id=id, base_url=base_url)
    rel_json = result["json_path"].relative_to(BASE_DIR).as_posix()
    result_url = f"{base_url.rstrip('/')}/{rel_json}"
    return {
        "id": id,
        "filename": pdf_name,
        "status": "completed",
        "result_url": result_url,
        "manifest": result["manifest"],
        "manifest_items": result["manifest"].get("items", []),
        "image_count": result["manifest"].get("image_count", 0),
    }

@router.post("/{user_id}")
async def upload_pdf_images(
    user_id: str = Path(..., description="User ID of the uploader"),
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    authorization: Optional[str] = Header(default=None, alias="Authorization"),
):
    try:
        base_url = os.getenv("BASE_URL", "http://localhost:8000")
        result = await handle_pdf_upload(user_id, file, background_tasks, authorization, base_url)
        if "error" in result:
            raise HTTPException(status_code=422, detail=result["error"])
        return JSONResponse(result)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{user_id}/update-visibility")
async def update_image_visibility(
    user_id: str = Path(..., description="User ID of the uploader"),
    body: Dict[str, Any] = None
):
    try:
        if not body or "pdf_stem" not in body:
            raise HTTPException(status_code=400, detail="Missing pdf_stem in body")
        pdf_stem = body["pdf_stem"]
        items_update = {it["file_name"]: bool(it.get("is_visible", True)) for it in body.get("items", [])}
        manifest_path = UPLOAD_ROOT / user_id / f"{pdf_stem}.json"
        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail="Manifest file not found")
        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)
        updated = 0
        for it in manifest.get("items", []):
            fname = it.get("file_name")
            if fname in items_update:
                it["is_visible"] = items_update[fname]
                updated += 1
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)
        return {
            "status": "updated",
            "updated_count": updated,
            "json_path": str(manifest_path),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class ProjectCreate(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    description: str = Field(..., min_length=1, max_length=2000)
    @validator("name", "description")
    def strip_ws(cls, v: str) -> str:
        s = v.strip()
        if not s:
            raise ValueError("must not be empty")
        return s

@router.get("/{user_id}/projects")
async def list_user_projects(
    user_id: int = Path(..., description="User ID"),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Project)
        .filter(Project.owner_id == user_id, Project.is_archived == False)
        .order_by(Project.updated_at.desc())
        .all()
    )
    data = [
        {"id": p.id, "name": p.name or "", "description": p.description or ""}
        for p in rows
    ]
    return {"projects": data}

@router.post("/{user_id}/projects", status_code=status.HTTP_201_CREATED)
async def create_user_project(
    payload: ProjectCreate,
    user_id: int = Path(..., description="User ID"),
    db: Session = Depends(get_db),
):
    existing = (
        db.query(Project)
        .filter(
            Project.owner_id == user_id,
            Project.is_archived == False,
            Project.name.ilike(payload.name.strip()),
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A project with this name already exists.")
    now = datetime.datetime.utcnow()
    proj = Project(
        owner_id=user_id,
        name=payload.name.strip(),
        description=payload.description.strip(),
        is_archived=False,
        created_at=now,
        updated_at=now,
    )
    db.add(proj)
    db.commit()
    db.refresh(proj)
    proj_dir = UPLOAD_ROOT / str(user_id) / "projects" / str(proj.id)
    ensure_dir(proj_dir)
    headers = {"Location": f"/uploads-pdf-images/{user_id}/projects/{proj.id}"}
    return JSONResponse(
        status_code=status.HTTP_201_CREATED,
        headers=headers,
        content={"id": proj.id, "name": proj.name, "description": proj.description},
    )

@router.post("/{user_id}/export")
async def export_alt_text_manifest(
    user_id: int = Path(..., description="User ID of the uploader"),
    body: Dict[str, Any] = None,
    db: Session = Depends(get_db),
):
    try:
        if not body or "pdf_stem" not in body or "format" not in body:
            raise HTTPException(status_code=400, detail="Missing pdf_stem or format")

        pdf_stem = body["pdf_stem"]
        fmt = str(body["format"]).lower()

        manifest_path = UPLOAD_ROOT / str(user_id) / f"{pdf_stem}.json"
        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail="Manifest file not found")

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        items = [
            it for it in manifest.get("items", [])
            if it.get("is_visible", True) and int(it.get("index", 0) or 0) > 0
        ]

        if not items:
            raise HTTPException(status_code=400, detail="No eligible items found (visible & index > 0)")

        project_id = body.get("project_id")

        if not project_id and body.get("project_name") and body.get("project_description"):
            now = datetime.datetime.utcnow()
            proj = Project(
                owner_id=user_id,
                name=str(body["project_name"]).strip(),
                description=str(body["project_description"]).strip(),
                is_archived=False,
                created_at=now,
                updated_at=now,
            )
            db.add(proj)
            db.commit()
            db.refresh(proj)
            project_id = proj.id

        if project_id:
            proj_row = (
                db.query(Project)
                .filter(
                    Project.id == project_id,
                    Project.owner_id == user_id,
                    Project.is_archived == False
                )
                .first()
            )

            if not proj_row:
                raise HTTPException(status_code=404, detail="Project not found")

        download_filename = ""
        media_type = ""
        tmp_path = None

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output)

            writer.writerow(["Page", "Index", "File Name", "Image"])

            for it in items:
                img_path = UPLOAD_ROOT / str(user_id) / pdf_stem / "images" / it.get("file_name")
                writer.writerow([
                    it.get("page"),
                    it.get("index"),
                    it.get("file_name"),
                    str(img_path)
                ])

            output.seek(0)

            data_str = output.getvalue()

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
            tmpf.write(data_str.encode("utf-8"))
            tmpf.flush()
            tmpf.close()

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}.csv"
            media_type = "text/csv"

        elif fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            MAX_ROWS_PER_SHEET = 100

            wb = Workbook()
            ws = wb.active
            ws.title = "Manifest 1"

            def setup_sheet(sheet):
                headers = ["Page", "Index", "File Name", "Image"]
                sheet.append(headers)
                sheet.column_dimensions[get_column_letter(1)].width = 8
                sheet.column_dimensions[get_column_letter(2)].width = 8
                sheet.column_dimensions[get_column_letter(3)].width = 35
                sheet.column_dimensions[get_column_letter(4)].width = 25

            setup_sheet(ws)

            current_row_count = 0
            sheet_index = 1

            for it in items:
                if current_row_count >= MAX_ROWS_PER_SHEET:
                    sheet_index += 1
                    ws = wb.create_sheet(title=f"Manifest {sheet_index}")
                    setup_sheet(ws)
                    current_row_count = 0

                page = it.get("page")
                index = it.get("index")
                fname = it.get("file_name")

                row = [page, index, fname, ""]
                ws.append(row)

                row_idx = ws.max_row

                img_path = UPLOAD_ROOT / str(user_id) / pdf_stem / "images" / fname

                if img_path.exists():
                    try:
                        img = XLImage(str(img_path))
                        img.width = 150
                        img.height = 150
                        ws.add_image(img, f"D{row_idx}")
                        ws.row_dimensions[row_idx].height = 120
                    except Exception:
                        pass

                for col_idx in range(1, 5):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        vertical="center",
                        horizontal="center"
                    )

                current_row_count += 1

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmpf.name)

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif fmt == "epub":
            images_dir = UPLOAD_ROOT / str(user_id) / pdf_stem / "images"

            if not images_dir.exists():
                raise HTTPException(status_code=404, detail="Images folder not found")

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".epub")

            with zipfile.ZipFile(tmpf.name, "w") as zf:
                zf.writestr("mimetype", "application/epub+zip", compress_type=zipfile.ZIP_STORED)

                zf.writestr(
                    "META-INF/container.xml",
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<container version="1.0" xmlns="urn:oasis:names:tc:opendocument:xmlns:container">'
                    '<rootfiles><rootfile full-path="OEBPS/content.opf" media-type="application/oebps-package+xml"/></rootfiles>'
                    "</container>",
                )

                book_id = str(uuid.uuid4())
                date_iso = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

                manifest_entries = []
                spine_entries = []
                nav_list_items = []

                for idx, it in enumerate(items, 1):
                    fname = it.get("file_name")
                    img_path = images_dir / fname

                    if not img_path.exists():
                        continue

                    img_id = f"img{idx}"
                    xhtml_id = f"item{idx}"

                    img_ext = img_path.suffix.lower().strip(".")
                    media = f"image/{'jpeg' if img_ext in ['jpg', 'jpeg'] else 'png'}"

                    with open(img_path, "rb") as fimg:
                        zf.writestr(
                            f"OEBPS/Images/{fname}",
                            fimg.read(),
                            compress_type=zipfile.ZIP_DEFLATED
                        )

                    page_html = (
                        '<?xml version="1.0" encoding="utf-8"?>'
                        '<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" lang="en">'
                        '<head><meta charset="utf-8"/><title>Image</title>'
                        '<style>body{margin:0;padding:0;text-align:center;}img{max-width:100%;height:auto;}</style>'
                        "</head><body>"
                        f'<figure><img src="Images/{fname}" alt="{fname}"/>'
                        f'<figcaption>Page {it.get("page")} · Index {it.get("index")} · {fname}</figcaption>'
                        "</figure></body></html>"
                    )

                    xhtml_path = f"OEBPS/Text/page_{idx}.xhtml"

                    zf.writestr(
                        xhtml_path,
                        page_html,
                        compress_type=zipfile.ZIP_DEFLATED
                    )

                    manifest_entries.append(f'<item id="{img_id}" href="Images/{fname}" media-type="{media}"/>')
                    manifest_entries.append(f'<item id="{xhtml_id}" href="Text/page_{idx}.xhtml" media-type="application/xhtml+xml"/>')

                    spine_entries.append(f'<itemref idref="{xhtml_id}"/>')

                    nav_list_items.append(f'<li><a href="Text/page_{idx}.xhtml">{fname}</a></li>')

                nav_html = (
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" '
                    'xmlns:epub="http://www.idpf.org/2007/ops" lang="en">'
                    "<head><meta charset=\"utf-8\"/><title>Contents</title></head>"
                    "<body><nav epub:type=\"toc\" id=\"toc\"><ol>"
                    + "".join(nav_list_items) +
                    "</ol></nav></body></html>"
                )

                zf.writestr("OEBPS/nav.xhtml", nav_html, compress_type=zipfile.ZIP_DEFLATED)

                manifest_entries.append(
                    '<item id="nav" href="nav.xhtml" media-type="application/xhtml+xml" properties="nav"/>'
                )

                content_opf = (
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<package version="3.0" xmlns="http://www.idpf.org/2007/opf" unique-identifier="bookid">'
                    f'<metadata xmlns:dc="http://purl.org/dc/elements/1.1/"><dc:identifier id="bookid">{book_id}</dc:identifier>'
                    f"<dc:title>{pdf_stem}</dc:title><dc:language>en</dc:language><dc:date>{date_iso}</dc:date></metadata>"
                    f"<manifest>{''.join(manifest_entries)}</manifest>"
                    f"<spine>{''.join(spine_entries)}</spine>"
                    "</package>"
                )

                zf.writestr("OEBPS/content.opf", content_opf, compress_type=zipfile.ZIP_DEFLATED)

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}.epub"
            media_type = "application/epub+zip"

        elif fmt == "mobi":
            images_dir = UPLOAD_ROOT / str(user_id) / pdf_stem / "images"

            if not images_dir.exists():
                raise HTTPException(status_code=404, detail="Images folder not found")

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".mobi.zip")

            with zipfile.ZipFile(tmpf.name, "w", zipfile.ZIP_DEFLATED) as zf:
                index_list = []

                for idx, it in enumerate(items, 1):
                    fname = it.get("file_name")
                    img_path = images_dir / fname

                    if not img_path.exists():
                        continue

                    with open(img_path, "rb") as fimg:
                        b64 = base64.b64encode(fimg.read()).decode("ascii")

                    mime = "image/jpeg" if fname.lower().endswith((".jpg", ".jpeg")) else "image/png"

                    html = (
                        "<!doctype html><html><head><meta charset=\"utf-8\"><title>Image</title>"
                        "<style>body{margin:0;padding:0;text-align:center;}img{max-width:100%;height:auto;}</style>"
                        "</head><body>"
                        f'<figure><img src="data:{mime};base64,{b64}" alt="{fname}"/>'
                        f'<figcaption>Page {it.get("page")} · Index {it.get("index")} · {fname}</figcaption>'
                        "</figure></body></html>"
                    )

                    html_name = f"page_{idx}.html"
                    zf.writestr(html_name, html)

                    index_list.append(f'<li><a href="{html_name}">{fname}</a></li>')

                toc = (
                    "<!doctype html><html><head><meta charset=\"utf-8\"><title>Contents</title></head>"
                    "<body><h1>Contents</h1><ol>" + "".join(index_list) + "</ol></body></html>"
                )

                zf.writestr("index.html", toc)

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}.mobi.zip"
            media_type = "application/zip"

        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {fmt}")

        base_url = "http://localhost:8000"
        export_dir = UPLOAD_ROOT / str(user_id) / pdf_stem

        ensure_dir(export_dir)

        final_path = export_dir / download_filename

        try:
            shutil.move(tmp_path, final_path)
        except Exception:
            shutil.copyfile(tmp_path, final_path)
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        public_url = f"{base_url}/uploads/{user_id}/{pdf_stem}/{download_filename}"

        if project_id:
            try:
                size = final_path.stat().st_size if final_path.exists() else 0
                ext = FilePath(download_filename).suffix.lower().lstrip(".")
                mime_guess, _ = mimetypes.guess_type(download_filename)

                now = datetime.datetime.utcnow()

                pf = ProjectFile(
                    project_id=project_id,
                    uploaded_by=user_id,
                    original_name=download_filename,
                    storage_path=public_url,
                    mime_type=mime_guess or media_type or "application/octet-stream",
                    ext=ext,
                    size_bytes=size,
                    checksum=None,
                    is_deleted=False,
                    created_at=now,
                    updated_at=now,
                )

                db.add(pf)
                db.commit()
                db.refresh(pf)

            except Exception:
                pass

        user_plan = (
            db.query(UserPlan)
            .filter(
                UserPlan.user_id == user_id,
                UserPlan.is_active == True
            )
            .first()
        )

        if user_plan:
            user_plan.pdf_used = (user_plan.pdf_used or 0) + 1
            db.commit()

        try:
            usage_log = UsageLog(
                user_id=user_id,
                user_plan_id=user_plan.id if user_plan else None,
                type="Pdf to Images Export",
                file_name=download_filename,
                credits_used=1,
                reference_id=project_id
            )

            db.add(usage_log)
            db.commit()
        except Exception:
            pass

        return FileResponse(
            str(final_path),
            media_type=media_type,
            filename=download_filename
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.post("/{user_id}/manual-crop")
async def save_manual_cropped_image(
    user_id: str = Path(..., description="User ID of the uploader"),
    pdf_stem: str = Form(...),
    page: int = Form(...),
    image: UploadFile = File(...)
):
    try:
        if page <= 0:
            raise HTTPException(status_code=400, detail="Invalid page number")

        user_dir = UPLOAD_ROOT / str(user_id)
        pdf_dir = user_dir / pdf_stem
        images_dir = pdf_dir / "images"
        manifest_path = user_dir / f"{pdf_stem}.json"

        if not pdf_dir.exists():
            raise HTTPException(status_code=404, detail="PDF folder not found")

        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail="Manifest file not found")

        ensure_dir(images_dir)

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        existing_indexes = []

        for item in manifest.get("items", []):
            try:
                item_page = int(item.get("page", 0) or 0)
                item_index = int(item.get("index", 0) or 0)

                if item_page == page and item_index > 0:
                    existing_indexes.append(item_index)
            except Exception:
                pass

        for existing_file in images_dir.glob(f"p{page}_img*.png"):
            try:
                name = existing_file.stem
                index_part = name.replace(f"p{page}_img", "")
                if index_part.isdigit():
                    existing_indexes.append(int(index_part))
            except Exception:
                pass

        next_index = max(existing_indexes) + 1 if existing_indexes else 1

        file_name = f"p{page}_img{next_index}.png"
        image_path = images_dir / file_name

        content = await image.read()

        if not content:
            raise HTTPException(status_code=400, detail="Empty image file")

        with image_path.open("wb") as f:
            f.write(content)

        base_url = os.getenv("BASE_URL", "http://localhost:8000").rstrip("/")

        new_item = {
            "page": page,
            "index": next_index,
            "type": "image",
            "has_text": False,
            "file_name": file_name,
            "system_path": str(image_path.resolve()),
            "web_path": f"{base_url}/uploads/{user_id}/{pdf_stem}/images/{file_name}",
            "is_visible": True
        }

        manifest.setdefault("items", []).append(new_item)
        manifest["image_count"] = int(manifest.get("image_count", 0) or 0) + 1

        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)

        return {
            "status": "saved",
            "message": "Cropped image saved successfully",
            "item": new_item,
            "image_count": manifest["image_count"]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))    
    
@router.post("/{user_id}/download-current")
async def download_current_manifest_file(
    user_id: int = Path(..., description="User ID of the uploader"),
    body: Dict[str, Any] = None
):
    try:
        if not body or "pdf_stem" not in body or "format" not in body:
            raise HTTPException(status_code=400, detail="Missing pdf_stem or format")

        pdf_stem = str(body["pdf_stem"]).strip()
        fmt = str(body["format"]).lower().strip()

        if fmt not in ["csv", "xlsx", "epub"]:
            raise HTTPException(status_code=400, detail="Supported formats are csv, xlsx, epub")

        user_dir = UPLOAD_ROOT / str(user_id)
        pdf_dir = user_dir / pdf_stem
        images_dir = pdf_dir / "images"
        manifest_path = user_dir / f"{pdf_stem}.json"

        if not pdf_dir.exists():
            raise HTTPException(status_code=404, detail="PDF folder not found")

        if not images_dir.exists():
            raise HTTPException(status_code=404, detail="Images folder not found")

        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail="Manifest file not found")

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        items = [
            it for it in manifest.get("items", [])
            if it.get("is_visible", True) and int(it.get("index", 0) or 0) > 0
        ]

        items = sorted(
            items,
            key=lambda it: (
                int(it.get("page", 0) or 0),
                int(it.get("index", 0) or 0)
            )
        )

        if not items:
            raise HTTPException(status_code=400, detail="No visible images found in current JSON")

        download_filename = ""
        media_type = ""
        tmp_path = None

        if fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output)

            writer.writerow(["Page", "Index", "File Name", "Image Path", "Image URL"])

            for it in items:
                fname = it.get("file_name")
                img_path = images_dir / fname

                if not fname or not img_path.exists():
                    continue

                writer.writerow([
                    it.get("page"),
                    it.get("index"),
                    fname,
                    str(img_path),
                    it.get("web_path", "")
                ])

            output.seek(0)

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
            tmpf.write(output.getvalue().encode("utf-8"))
            tmpf.flush()
            tmpf.close()

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}_current.csv"
            media_type = "text/csv"

        elif fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

            MAX_ROWS_PER_SHEET = 100

            wb = Workbook()
            ws = wb.active
            ws.title = "Images 1"

            header_fill = PatternFill("solid", fgColor="F1F5F9")
            header_font = Font(bold=True, color="0F172A")
            thin_border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0")
            )

            def setup_sheet(sheet):
                headers = ["Page", "Index", "File Name", "Image"]
                sheet.append(headers)

                sheet.column_dimensions[get_column_letter(1)].width = 10
                sheet.column_dimensions[get_column_letter(2)].width = 10
                sheet.column_dimensions[get_column_letter(3)].width = 38
                sheet.column_dimensions[get_column_letter(4)].width = 28

                for col_idx in range(1, 5):
                    cell = sheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", horizontal="center")

                sheet.row_dimensions[1].height = 24

            setup_sheet(ws)

            current_row_count = 0
            sheet_index = 1

            for it in items:
                if current_row_count >= MAX_ROWS_PER_SHEET:
                    sheet_index += 1
                    ws = wb.create_sheet(title=f"Images {sheet_index}")
                    setup_sheet(ws)
                    current_row_count = 0

                page = it.get("page")
                index = it.get("index")
                fname = it.get("file_name")

                if not fname:
                    continue

                img_path = images_dir / fname

                if not img_path.exists():
                    continue

                ws.append([page, index, fname, ""])
                row_idx = ws.max_row

                try:
                    img = XLImage(str(img_path))
                    img.width = 150
                    img.height = 150
                    ws.add_image(img, f"D{row_idx}")
                    ws.row_dimensions[row_idx].height = 120
                except Exception:
                    pass

                for col_idx in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

                current_row_count += 1

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmpf.name)

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}_current.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif fmt == "epub":
            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".epub")

            with zipfile.ZipFile(tmpf.name, "w") as zf:
                zf.writestr("mimetype", "application/epub+zip", compress_type=zipfile.ZIP_STORED)

                zf.writestr(
                    "META-INF/container.xml",
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<container version="1.0" xmlns="urn:oasis:names:tc:opendocument:xmlns:container">'
                    '<rootfiles><rootfile full-path="OEBPS/content.opf" media-type="application/oebps-package+xml"/></rootfiles>'
                    "</container>",
                )

                book_id = str(uuid.uuid4())
                date_iso = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

                manifest_entries = []
                spine_entries = []
                nav_list_items = []

                valid_index = 0

                for it in items:
                    fname = it.get("file_name")

                    if not fname:
                        continue

                    img_path = images_dir / fname

                    if not img_path.exists():
                        continue

                    valid_index += 1

                    img_id = f"img{valid_index}"
                    xhtml_id = f"item{valid_index}"

                    img_ext = img_path.suffix.lower().strip(".")
                    media = f"image/{'jpeg' if img_ext in ['jpg', 'jpeg'] else 'png'}"

                    with open(img_path, "rb") as fimg:
                        zf.writestr(
                            f"OEBPS/Images/{fname}",
                            fimg.read(),
                            compress_type=zipfile.ZIP_DEFLATED
                        )

                    page_html = (
                        '<?xml version="1.0" encoding="utf-8"?>'
                        '<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" lang="en">'
                        '<head><meta charset="utf-8"/><title>Image</title>'
                        '<style>'
                        'body{margin:0;padding:20px;text-align:center;font-family:Arial,sans-serif;}'
                        'img{max-width:100%;height:auto;}'
                        'figcaption{margin-top:12px;font-size:14px;color:#334155;}'
                        '</style>'
                        "</head><body>"
                        f'<figure><img src="Images/{fname}" alt="{fname}"/>'
                        f'<figcaption>Page {it.get("page")} · Image {it.get("index")} · {fname}</figcaption>'
                        "</figure></body></html>"
                    )

                    xhtml_path = f"OEBPS/Text/page_{valid_index}.xhtml"

                    zf.writestr(
                        xhtml_path,
                        page_html,
                        compress_type=zipfile.ZIP_DEFLATED
                    )

                    manifest_entries.append(f'<item id="{img_id}" href="Images/{fname}" media-type="{media}"/>')
                    manifest_entries.append(f'<item id="{xhtml_id}" href="Text/page_{valid_index}.xhtml" media-type="application/xhtml+xml"/>')

                    spine_entries.append(f'<itemref idref="{xhtml_id}"/>')
                    nav_list_items.append(f'<li><a href="Text/page_{valid_index}.xhtml">{fname}</a></li>')

                if valid_index == 0:
                    raise HTTPException(status_code=400, detail="No valid image files found for EPUB")

                nav_html = (
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" '
                    'xmlns:epub="http://www.idpf.org/2007/ops" lang="en">'
                    '<head><meta charset="utf-8"/><title>Contents</title></head>'
                    "<body><nav epub:type=\"toc\" id=\"toc\"><ol>"
                    + "".join(nav_list_items) +
                    "</ol></nav></body></html>"
                )

                zf.writestr("OEBPS/nav.xhtml", nav_html, compress_type=zipfile.ZIP_DEFLATED)

                manifest_entries.append(
                    '<item id="nav" href="nav.xhtml" media-type="application/xhtml+xml" properties="nav"/>'
                )

                content_opf = (
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<package version="3.0" xmlns="http://www.idpf.org/2007/opf" unique-identifier="bookid">'
                    f'<metadata xmlns:dc="http://purl.org/dc/elements/1.1/">'
                    f'<dc:identifier id="bookid">{book_id}</dc:identifier>'
                    f"<dc:title>{pdf_stem}</dc:title>"
                    "<dc:language>en</dc:language>"
                    f"<dc:date>{date_iso}</dc:date>"
                    "</metadata>"
                    f"<manifest>{''.join(manifest_entries)}</manifest>"
                    f"<spine>{''.join(spine_entries)}</spine>"
                    "</package>"
                )

                zf.writestr("OEBPS/content.opf", content_opf, compress_type=zipfile.ZIP_DEFLATED)

            tmp_path = tmpf.name
            download_filename = f"{pdf_stem}_current.epub"
            media_type = "application/epub+zip"

        final_path = pdf_dir / download_filename

        try:
            if final_path.exists():
                final_path.unlink()
        except Exception:
            pass

        try:
            shutil.move(tmp_path, final_path)
        except Exception:
            shutil.copyfile(tmp_path, final_path)
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        return FileResponse(
            str(final_path),
            media_type=media_type,
            filename=download_filename
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))    