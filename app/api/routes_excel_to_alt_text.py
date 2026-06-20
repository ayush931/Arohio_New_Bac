# app/api/v1/uploads_excel_to_alttext.py
from fastapi import APIRouter, UploadFile, File, Path, HTTPException, Body, Depends
from fastapi.responses import JSONResponse, FileResponse
from pathlib import Path as FilePath
from typing import Dict, Any, List, Optional, Tuple

import os, io, re, csv, json, time, math, uuid, shutil, base64, zipfile, mimetypes, tempfile
import datetime
import requests
import pandas as pd
from PIL import Image
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor, AbsoluteAnchor

from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.project import Project
from app.models.UserPlan import UserPlan
from app.models.UsageLog import UsageLog
import logging
from sqlalchemy import text
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)
try:
    from app.models.project_file import ProjectFile 
except Exception:
    ProjectFile = None  

router = APIRouter(prefix="/uploads-excel-to-alttext", tags=["Excel → AltText"])

BASE_DIR = FilePath(__file__).resolve().parents[2] if "__file__" in globals() else FilePath(".")
UPLOAD_ROOT = BASE_DIR / "uploads"
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)


PUBLIC_BASE_URL = os.getenv("BASE_URL", "http://localhost:8000")

def ensure_dir(p: FilePath):
    p.mkdir(parents=True, exist_ok=True)

# ----------------------------- Helpers for reading embedded images from Excel -----------------------------
EMU_PER_POINT = 12700

def _default_row_height_points(ws):
    try:
        h = ws.sheet_format.defaultRowHeight
        return float(h) if h else 15.0
    except:
        return 15.0

def _row_height_points(ws, r):
    try:
        h = ws.row_dimensions[r].height
        return float(h) if h else _default_row_height_points(ws)
    except:
        return _default_row_height_points(ws)

def _anchor_row(ws, anchor, data_start_row=2):
    if isinstance(anchor, str):
        r, _ = coordinate_to_tuple(anchor)
        return max(r, data_start_row)
    if isinstance(anchor, AbsoluteAnchor):
        return None
    if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
        base_r = int(anchor._from.row) + 1
        off = int(getattr(anchor._from, "rowOff", 0) or 0)
        if off <= 0:
            return max(base_r, data_start_row)
        y = off
        r = base_r
        while y > 0 and r <= ws.max_row + 1000:
            h_pt = _row_height_points(ws, r)
            h_emu = int(h_pt * EMU_PER_POINT)
            if y < h_emu:
                break
            y -= h_emu
            r += 1
        return max(r, data_start_row)
    return None

def _extract_embedded_images(ws) -> Dict[int, List[bytes]]:
    row_to_blobs: Dict[int, List[bytes]] = {}
    imgs = getattr(ws, "_images", []) or []
    for im in imgs:
        try:
            anch = getattr(im, "anchor", None)
            r = _anchor_row(ws, anch, data_start_row=2)
            if not r:
                continue
            raw = None
            try:
                raw = im._data() if hasattr(im, "_data") else None
            except:
                raw = None
            if raw is None:
                try:
                    raw = getattr(getattr(im, "ref", None), "blob", None)
                except:
                    raw = None
            if not raw:
                continue
            row_to_blobs.setdefault(r, []).append(raw)
        except:
            continue
    return row_to_blobs

# -------------------------------- LLaVA / Alt-text Prompting --------------------------------
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://127.0.0.1:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llava:latest")

ALT_RULES = (
    "ROLE: You are an expert accessibility writer for screen-reader users. "
    "Write accurate, useful, detailed alternative text that describes only what is visible in the attached visual.\n"
    "\n"
    "GLOBAL RULES:\n"
    "- Describe only what is actually visible. Do not invent or assume hidden information.\n"
    "- Do not guess brand, identity, location, purpose, medical use, product use, or meaning unless it is clearly visible.\n"
    "- Prioritise clearly readable visible text. If text is readable, include it exactly in the long alt.\n"
    "- If text is partly readable, include only the readable words and mention that some text is unreadable.\n"
    "- If text is present but not readable, say that text is present but unreadable. Do not invent it.\n"
    "- Do not copy any sample or example text from the prompt.\n"
    "- Do not use subjective words such as beautiful, modern, professional, premium, attractive, or high quality.\n"
    "- Do not add conclusions, claims, warnings, explanations, or interpretation beyond what is visible.\n"
    "- Use UK English and present tense.\n"
    "- Short alt and long alt must be different.\n"
    "- If the visual is purely decorative and has no useful meaning, return empty strings for both fields.\n"
    "- Output JSON ONLY with exactly the keys: short_alt, long_alt. No extra text.\n"
    "\n"
    "VISUAL TYPE RULES:\n"
    "- Identify the visual as photo, screenshot, diagram, flow chart, chart, graph, map, logo, poster, book cover, document, table, chemical structure, equation, illustration, or other accurate visible type.\n"
    "- Use photo for real-world objects, machinery, people, products, packaging, rooms, vehicles, equipment, buildings, food, medicine bottles, or physical scenes.\n"
    "- Use screenshot for software screens, websites, dashboards, mobile apps, forms, browser windows, or UI layouts.\n"
    "- Use chart or graph for bars, lines, pies, plotted data, axes, legends, or visualised values.\n"
    "- Use diagram or flow chart for labelled boxes, arrows, process steps, architecture, systems, relationships, or instructional visuals.\n"
    "- Use table for rows and columns of structured data.\n"
    "- Use document or page for scanned pages, forms, letters, certificates, book pages, or text-heavy pages.\n"
    "- Use logo only when the main content is a logo or brand mark.\n"
    "- Never call a real-world product, machinery, or medicine photo a book cover unless it is actually a book cover.\n"
    "\n"
    "SHORT ALT RULES:\n"
    "- One sentence only.\n"
    "- Maximum 190 characters.\n"
    "- Start with 'A' or 'An'.\n"
    "- Include the visual type explicitly.\n"
    "- Describe the main visible subject first.\n"
    "- Include the most important readable text only if it fits naturally.\n"
    "- Keep it concise but specific.\n"
    "- End with one full stop.\n"
    "\n"
    "LONG ALT RULES:\n"
    "- One detailed paragraph only.\n"
    "- Start with 'The'.\n"
    "- Include the exact visual type near the beginning.\n"
    "- Describe the main subject, layout, important objects, visible text, and relevant relationships between elements.\n"
    "- Mention positions only when useful using: top left, top centre, top right, centre left, centre, centre right, bottom left, bottom centre, bottom right.\n"
    "- End with one full stop.\n"
    "\n"
    "PRODUCT AND PACKAGING PHOTOS:\n"
    "- Describe visible boxes, bottles, labels, packets, containers, seals, and arrangement.\n"
    "- Transcribe readable product names, label text, numbers, strengths, quantities, company names, and visible warnings exactly as shown.\n"
    "- If multiple products are visible, describe them left to right.\n"
    "- Do not state what the product treats or does unless that exact information is visible.\n"
    "\n"
    "MACHINERY AND EQUIPMENT PHOTOS:\n"
    "- Describe visible machinery, control panels, buttons, screens, conveyors, guards, rollers, pipes, bottles, containers, labels, signs, and surrounding equipment.\n"
    "- Mention readable panel labels, warning signs, display text, numbers, or stickers if visible.\n"
    "- Do not assume the machine type, industry, or process unless clearly visible.\n"
    "\n"
    "PEOPLE AND SCENE PHOTOS:\n"
    "- Describe people by visible action, clothing, position, and context without identifying them.\n"
    "- Do not guess age, ethnicity, profession, emotion, relationship, or intent unless clearly visible.\n"
    "- Describe important objects, setting, and visible text.\n"
    "\n"
    "SCREENSHOTS AND UI:\n"
    "- Describe the screen type, main heading, fields, buttons, tabs, menus, cards, tables, alerts, and visible messages.\n"
    "- Transcribe important UI text exactly in logical reading order.\n"
    "- Mention selected states, error messages, success messages, or disabled controls if visible.\n"
    "\n"
    "CHARTS AND GRAPHS:\n"
    "- Describe the chart or graph type first.\n"
    "- Include title, axis labels, units, legend items, categories, visible values, and notable visible trends only when readable.\n"
    "- Do not calculate or infer values that are not visible.\n"
    "- If exact values are unreadable, describe the visible pattern generally.\n"
    "\n"
    "DIAGRAMS AND FLOW CHARTS:\n"
    "- Describe boxes, nodes, arrows, labels, connectors, direction of flow, hierarchy, and grouped sections.\n"
    "- Transcribe all readable labels in logical order.\n"
    "- Summarise relationships between elements without adding hidden context.\n"
    "\n"
    "TABLES AND DOCUMENTS:\n"
    "- Describe the document or table type, headings, rows, columns, sections, and important readable text.\n"
    "- Preserve visible order and wording as much as possible.\n"
    "- If the page is text-heavy, summarise structure and include the most important readable text.\n"
    "\n"
    "MAPS:\n"
    "- Describe the map area, visible labels, markers, routes, legends, scale, compass, and important regions if readable.\n"
    "- Do not invent place names or distances.\n"
    "\n"
    "LOGOS AND POSTERS:\n"
    "- For logos, describe the visible mark and transcribe logotype text if readable. Do not guess the brand if text is unreadable.\n"
    "- For posters, include heading, date, venue, people, objects, call-to-action text, QR codes, and sponsor text if visible.\n"
    "\n"
    "CHEMICAL STRUCTURES, EQUATIONS, AND SCIENTIFIC VISUALS:\n"
    "- Describe molecular structures, labels, equations, symbols, arrows, axes, and annotations that are readable.\n"
    "- Convert mathematical notation into words when needed for screen-reader clarity.\n"
    "- Do not infer scientific meaning beyond visible labels and structure.\n"
)
ALT_FEWSHOTS = """
Output strict JSON only with exactly these keys: short_alt, long_alt.

Describe only what is visible.
Do not invent text, labels, medicine names, product names, company names, chart titles, dates, numbers, or values.
If readable text exists, include it in the long alt.
If text is partly readable, include readable words and say other text is unreadable.
If text is present but unreadable, say text is present but unreadable.
Short alt must be concise.
Long alt must be detailed, factual, and useful for screen-reader users.
"""

BANNED_SYMBOLS = set('%/–:;^_°"“”‘’\'')
VT_WORDS = ("photo","diagram","illustration","flow chart","logo","map","screenshot","chemical structure","chart","graph","poster","book cover","cover")

def _strip_banned_symbols(t: str) -> str:
    return "".join(ch for ch in t if ch not in BANNED_SYMBOLS)

def _ensure_final_period(t: str) -> str:
    t = t.strip()
    if t and not t.endswith("."):
        t += "."
    return t

def _normalise_prefix_spaces(t: str) -> str:
    return re.sub(r"\s+", " ", t).strip()

def _remove_banned_words_anywhere(t: str) -> str:
    return re.sub(r"\b(image|figure|picture)s?\b", "visual", t, flags=re.I)

def _force_visual_type_prefix(text: str, vtype: str, short: bool) -> str:
    t = _normalise_prefix_spaces(text)
    t = _remove_banned_words_anywhere(t)
    t = re.sub(r"^(The|A|An)\s+photo\s+(A|An|The)\s+", r"\1 ", t, flags=re.I)
    t = re.sub(r"^(The|A|An)\s+visual\s+(A|An|The)\s+", r"\1 ", t, flags=re.I)
    t = re.sub(r"^(A|An|The)\s+(" + "|".join(re.escape(v) for v in VT_WORDS) + r")\s+", "", t, flags=re.I)
    if short:
        article = "An" if vtype[:1].lower() in "aeiou" else "A"
        t = f"{article} {vtype} {t}".strip()
    else:
        t = f"The {vtype} {t}".strip()
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _truncate_words(text: str, limit: Optional[int]) -> str:
    if not limit or limit <= 0:
        return text
    if len(text) <= limit:
        return text
    cut = text[:limit].rstrip()
    if " " in cut:
        cut = cut[:cut.rfind(" ")].rstrip()
    return cut

def choose_visual_type(item: Dict[str, Any]) -> str:
    name = (item.get("file_name") or item.get("File Name") or "").lower()
    path = ((item.get("image_path") or "") + " " + " ".join(item.get("image_paths", []))).lower()
    hay = f"{name} {path}"

    if any(h in hay for h in ("cvr", "cover", "cvr_hr", "front_cover", "book_cover")):
        return "book cover"
    if "screenshot" in hay or "login" in hay or "ui_" in hay:
        return "screenshot"
    if "logo" in hay:
        return "logo"
    if "poster" in hay:
        return "poster"
    if "chart" in hay or "graph" in hay:
        return "chart"
    if "diagram" in hay or "flow" in hay:
        return "diagram"
    if "map" in hay:
        return "map"

    return "photo"

def build_llava_prompt(vtype: str, visible_text: str) -> str:
    vt_text_block = "None" if not (visible_text or "").strip() else visible_text.strip()

    return (
        "You are an accessibility alt text writer for screen-reader users.\n"
        "Describe only the attached visual. Do not invent hidden details.\n"
        "Return strict JSON only with keys short_alt and long_alt.\n"
        "\n"
        "Identify the correct visual type yourself.\n"
        "Use chart or graph for axes, plotted lines, bars, legends, values, or numeric data.\n"
        "Use photo for real-world objects, machinery, products, packaging, bottles, people, rooms, equipment, or physical scenes.\n"
        "Use screenshot for software screens, websites, dashboards, forms, browser windows, or app UI.\n"
        "Use diagram or flow chart for boxes, arrows, nodes, process steps, or labelled relationships.\n"
        "Use table for rows and columns. Use document for text-heavy pages or forms.\n"
        "\n"
        "Short alt rules:\n"
        "- One sentence only.\n"
        "- Start with A or An.\n"
        "- Include the correct visual type.\n"
        "- Mention the main visible subject and the most important readable text if useful.\n"
        "- Maximum 190 characters.\n"
        "\n"
        "Long alt rules:\n"
        "- One detailed paragraph only.\n"
        "- Start with The.\n"
        "- Include the correct visual type near the beginning.\n"
        "- Describe layout, main subject, important objects, readable text, and relationships between elements.\n"
        "- For machinery, mention controls, panels, buttons, displays, conveyors, guards, containers, labels, signs, and surrounding equipment.\n"
        "- For products or medicine packaging, mention visible boxes, bottles, labels, product names, numbers, strengths, quantities, and company names exactly if readable.\n"
        "- For charts or graphs, mention chart type, axes, scale, legends, visible values, labels, and visible trend only if readable. Do not invent title or values.\n"
        "- For screenshots, mention headings, fields, buttons, tabs, menus, alerts, tables, and visible messages.\n"
        "- For diagrams, mention boxes, arrows, labels, flow direction, and grouped sections.\n"
        "- If text is partly readable, include readable words and say other text is unreadable.\n"
        "- If text is present but unreadable, say text is present but unreadable.\n"
        "- Do not add assumptions, claims, medical purpose, interpretation, or brand guesses.\n"
        "\n"
        f"System suggested type: {vtype}\n"
        f"Visible text hint: {vt_text_block}\n"
        "\n"
        "JSON format:\n"
        "{\"short_alt\":\"...\",\"long_alt\":\"...\"}"
    )


def _extract_json_object(txt: str) -> Optional[Dict[str, Any]]:
    i = 0
    n = len(txt)
    while i < n and txt[i] != '{':
        i += 1
    if i >= n:
        return None
    depth = 0
    in_str = False
    esc = False
    start = i
    while i < n:
        ch = txt[i]
        if in_str:
            if esc:
                esc = False
            elif ch == '\\':
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    block = txt[start:i+1]
                    try:
                        return json.loads(block)
                    except Exception:
                        j = i + 1
                        while j < n and txt[j] != '{':
                            j += 1
                        if j < n:
                            i = j - 1
                            start = j
                            depth = 0
                            in_str = False
                            esc = False
                        else:
                            return None
        i += 1
    return None

def clean_and_validate_alt(short_alt: str, long_alt: str, vtype: str, short_limit: Optional[int], long_limit: Optional[int]) -> Tuple[str, str]:
    s = _strip_banned_symbols(short_alt or "")
    s = _force_visual_type_prefix(s, vtype, short=True)
    s = _truncate_words(s, short_limit if short_limit else 190).strip()
    s = _ensure_final_period(s)

    l = (long_alt or "")
    l = re.sub(r"\s+", " ", l).strip()
    l = _force_visual_type_prefix(l, vtype, short=False)
    l = _truncate_words(l, long_limit if long_limit else None)
    l = _ensure_final_period(l)

    if s.strip(".").strip().lower() == l.strip(".").strip().lower():
        l = l.rstrip(".") + " It lists all visible text in reading order."
    l = _ensure_final_period(l)
    return s, l

def llava_alt_text(img_path: str, vtype: str, visible_text: str, short_limit: Optional[int], long_limit: Optional[int]) -> Tuple[str, str]:
    default_s = f"A {vtype} showing the main subject and prominent text."
    default_l = f"The {vtype} shows the main subject and visible text in reading order."

    try:
        logger.info(f"Starting LLaVA alt text generation for image: {img_path}")
        logger.info(f"OLLAMA_HOST: {OLLAMA_HOST}")
        logger.info(f"OLLAMA_MODEL: {OLLAMA_MODEL}")

        if not FilePath(img_path).exists():
            logger.error(f"Image path does not exist: {img_path}")
            return clean_and_validate_alt(default_s, default_l, vtype, short_limit, long_limit)

        with open(img_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")

        prompt = build_llava_prompt(vtype, visible_text or "")

        payload = {
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "images": [b64],
            "stream": False,
            "options": {
                "temperature": 0.1,
                "top_p": 0.3,
                "repeat_penalty": 1.07,
                "num_predict": 500
            },
        }

        r = requests.post(f"{OLLAMA_HOST.rstrip('/')}/api/generate", json=payload, timeout=180)

        logger.info(f"Ollama status code: {r.status_code}")
        logger.info(f"Ollama raw response: {r.text[:2000]}")

        r.raise_for_status()

        response_json = r.json()
        txt = response_json.get("response", "").strip()

        logger.info(f"Ollama parsed response for {img_path}: {txt}")

        s, l = default_s, default_l

        data = _extract_json_object(txt)

        if data is None and "{" in txt and "}" in txt:
            js = txt[txt.find("{"):txt.rfind("}") + 1]
            try:
                data = json.loads(js)
            except Exception as parse_error:
                logger.error(f"JSON parse failed from Ollama response: {str(parse_error)}")
                data = None

        if isinstance(data, dict):
            s = str(data.get("short_alt", s)).strip()
            l = str(data.get("long_alt", l)).strip()
        else:
            logger.error(f"No valid JSON object found in Ollama response for image: {img_path}")

        return clean_and_validate_alt(s, l, vtype, short_limit, long_limit)

    except Exception as e:
        logger.error(f"LLaVA alt text generation failed for {img_path}: {str(e)}", exc_info=True)
        return clean_and_validate_alt(default_s, default_l, vtype, short_limit, long_limit)
# -------------------------------- Generate (batch) --------------------------------
@router.post("/generate")
async def generate_alt_text(payload: dict = Body(...)):
    try:
        user_id = payload.get("user_id")
        items = payload.get("items", [])
        short_limit = payload.get("short_limit", None)
        long_limit = payload.get("long_limit", None)

        if not user_id:
            raise HTTPException(status_code=400, detail="Missing user_id")
        if not isinstance(items, list):
            raise HTTPException(status_code=400, detail="items must be a list")

        excel_stem = payload.get("excel_stem", "")
        if not excel_stem:
            src_url = payload.get("source_excel_url", "")
            if not src_url:
                raise HTTPException(status_code=400, detail="Provide excel_stem or source_excel_url to locate the manifest")
            try:
                excel_stem = FilePath(src_url).parent.name
            except Exception:
                raise HTTPException(status_code=400, detail="Unable to infer excel_stem from source_excel_url")

        user_dir = UPLOAD_ROOT / str(user_id)
        manifest_path = user_dir / f"{excel_stem}.json"
        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail=f"Manifest not found at {manifest_path}")

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        if short_limit is not None:
            manifest["short_limit"] = short_limit
        if long_limit is not None:
            manifest["long_limit"] = long_limit

        existing_items = manifest.get("items", [])
        by_row = {it.get("row"): it for it in existing_items if it.get("row") is not None}
        by_index = {it.get("index"): it for it in existing_items if it.get("index") is not None}
        by_name = {it.get("file_name"): it for it in existing_items if it.get("file_name")}

        updated = 0
        for patch in items:
            target = None
            r = patch.get("row", None)
            if r is not None and r in by_row:
                target = by_row[r]
            elif patch.get("index") in by_index:
                target = by_index.get(patch.get("index"))
            elif patch.get("file_name") in by_name:
                target = by_name.get(patch.get("file_name"))
            if not target:
                continue

            if "is_visible" in patch:
                target["is_visible"] = bool(patch["is_visible"])
                if not target["is_visible"]:
                    target.pop("short_alt_text", None)
                    target.pop("long_alt_text", None)
                    target.pop("visual_type", None)
                    target.pop("alt_generated_at", None)

            if "image_url" in patch:
                target["image_url"] = patch["image_url"]
            if "image_urls" in patch and isinstance(patch["image_urls"], list):
                target["image_urls"] = patch["image_urls"]
            if "image_paths" in patch and isinstance(patch["image_paths"], list):
                target["image_paths"] = patch["image_paths"]
            updated += 1

        to_generate = [it for it in existing_items if it.get("is_visible", True)]
        generated = 0
        for it in to_generate:
            img_path = None
            paths = it.get("image_paths") or []
            for p in paths:
                if p and FilePath(p).exists():
                    img_path = p
                    break
            if not img_path:
                sp = it.get("image_path") or it.get("system_path")
                if sp and FilePath(sp).exists():
                    img_path = sp
            if not img_path:
                continue

            vtype = choose_visual_type(it)
            vis_text = it.get("visible_text", "")
            s, l = llava_alt_text(img_path, vtype, vis_text, short_limit, long_limit)
            it["short_alt_text"] = s
            it["long_alt_text"] = l
            it["short_alt"] = s
            it["long_alt"] = l
            it["visual_type"] = vtype
            it["alt_generated_at"] = datetime.datetime.utcnow().isoformat() + "Z"
            generated += 1

        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)

        return JSONResponse({
    "status": "ok",
    "message": "Manifest updated",
    "saved_path": str(manifest_path.resolve()),
    "excel_stem": excel_stem,
    "updated_items": updated,
    "generated_items": generated,
    "short_limit": manifest.get("short_limit"),
    "long_limit": manifest.get("long_limit"),
    "row_count": len(manifest.get("items", [])),
    "items": manifest.get("items", [])
})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# -------------------------------- Regenerate (single item) --------------------------------
@router.post("/regenerate")
async def regenerate_alt_text(payload: dict = Body(...)):
    """
    Regenerate alt text for a single item and persist into the manifest JSON.
    Expected body:
    {
      "user_id": "123",
      "excel_stem": "my_excel",
      "page": 1,            # optional; used if present on items
      "index": 3,           # 1-based
      "short_limit": 125,   # optional
      "long_limit": 250     # optional
    }
    """
    try:
        user_id = payload.get("user_id")
        excel_stem = payload.get("excel_stem")
        page = payload.get("page", None)
        idx = payload.get("index", None)
        short_limit = payload.get("short_limit", None)
        long_limit = payload.get("long_limit", None)

        if not user_id:
            raise HTTPException(status_code=400, detail="Missing user_id")
        if not excel_stem:
            raise HTTPException(status_code=400, detail="Missing excel_stem")
        if idx is None:
            raise HTTPException(status_code=400, detail="Missing index")

        user_dir = UPLOAD_ROOT / str(user_id)
        manifest_path = user_dir / f"{excel_stem}.json"
        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail=f"Manifest not found at {manifest_path}")

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)

        items = manifest.get("items", [])
        if not isinstance(items, list) or not items:
            raise HTTPException(status_code=404, detail="No items in manifest")

        target = None
        if page is not None:
            for it in items:
                if int(it.get("index", -9999)) == int(idx) and int(it.get("page", -9999)) == int(page):
                    target = it
                    break
        if target is None:
            for it in items:
                if int(it.get("index", -9999)) == int(idx):
                    target = it
                    break
        if target is None:
            for it in items:
                if int(it.get("row", -9999)) == int(idx):
                    target = it
                    break
        if target is None:
            raise HTTPException(status_code=404, detail=f"Item not found for index={idx} (page={page})")

        img_path = None
        paths = target.get("image_paths") or []
        for p in paths:
            if p and FilePath(p).exists():
                img_path = p
                break
        if not img_path:
            sp = target.get("image_path") or target.get("system_path")
            if sp and FilePath(sp).exists():
                img_path = sp
        if not img_path:
            raise HTTPException(status_code=404, detail="No valid image path for this item")

        vtype = choose_visual_type(target)
        vis_text = target.get("visible_text", "")
        short_alt, long_alt = llava_alt_text(img_path, vtype, vis_text, short_limit, long_limit)

        target["short_alt_text"] = short_alt
        target["long_alt_text"] = long_alt
        target["short_alt"] = short_alt
        target["long_alt"] = long_alt
        target["visual_type"] = vtype
        target["alt_generated_at"] = datetime.datetime.utcnow().isoformat() + "Z"

        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, indent=2)

        return JSONResponse({
            "status": "ok",
            "message": "Regenerated and saved",
            "excel_stem": excel_stem,
            "saved_path": str(manifest_path.resolve()),
            "page": target.get("page"),
            "index": target.get("index"),
            "file_name": target.get("file_name"),
            "image_url": target.get("image_url") or (target.get("image_urls", [None]) or [None])[0],
            "short_alt": short_alt,
            "long_alt": long_alt,
            "visual_type": vtype,
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# -------------------------------- Upload Excel/CSV → manifest --------------------------------
@router.post("/file/{user_id}")
async def upload_excel_to_alttext(
    user_id: str = Path(..., description="User ID of the uploader"),
    file: UploadFile = File(...),
):
    try:
        fname = file.filename or "upload.xlsx"
        ext = fname.lower().rsplit(".", 1)[-1]
        if ext not in ["xlsx", "xls", "csv"]:
            raise HTTPException(status_code=400, detail="Only Excel (.xlsx/.xls) or CSV files allowed")

        excel_stem = FilePath(fname).stem
        user_dir = UPLOAD_ROOT / str(user_id)
        excel_dir = user_dir / excel_stem
        ensure_dir(excel_dir)
        excel_path = excel_dir / fname
        with excel_path.open("wb") as f:
            shutil.copyfileobj(file.file, f)

        images_dir = excel_dir / "images"
        ensure_dir(images_dir)

        items: List[Dict[str, Any]] = []

        if ext in ["xlsx", "xls"]:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

            row_to_blobs = _extract_embedded_images(ws)

            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
                row_map: Dict[str, Any] = {}
                for h, c in zip(headers, row):
                    key = h if h else ""
                    row_map[key] = "" if c.value is None else str(c.value).strip()

                excel_row_idx = i + 1
                blobs = row_to_blobs.get(excel_row_idx, [])
                saved_files = []
                for j, b in enumerate(blobs, start=1):
                    try:
                        pil = Image.open(BytesIO(b))
                        fmt = (pil.format or "PNG").upper()
                        ext2 = ".png" if fmt not in ["JPEG", "JPG", "PNG", "GIF", "BMP", "WEBP"] else f".{fmt.lower() if fmt!='JPG' else 'jpg'}"
                        name = f"r{excel_row_idx}_img{j}{ext2}"
                        outp = images_dir / name
                        if ext2 in [".jpg", ".jpeg"]:
                            pil.convert("RGB").save(outp, quality=92)
                        else:
                            pil.save(outp)
                        saved_files.append({
                            "file_name": name,
                            "system_path": str(outp.resolve()),
                            "web_path": f"{PUBLIC_BASE_URL}/uploads/{user_id}/{excel_stem}/images/{name}",
                        })
                    except:
                        continue

                row_map["image_url"]  = saved_files[0]["web_path"] if saved_files else ""
                row_map["image_urls"] = [x["web_path"] for x in saved_files]
                row_map["image_paths"] = [x["system_path"] for x in saved_files]
                row_map["row"] = i
                row_map["index"] = i
                row_map["is_visible"] = True
                items.append(row_map)
            wb.close()

        else:
            df = pd.read_csv(excel_path)
            for i, r in df.iterrows():
                row_map = {str(c): ("" if pd.isna(r[c]) else str(r[c])) for c in df.columns}
                row_map["image_url"] = ""
                row_map["image_urls"] = []
                row_map["image_paths"] = []
                row_map["row"] = i + 1
                row_map["index"] = i + 1
                row_map["is_visible"] = True
                items.append(row_map)

        manifest = {
            "source_excel": str(excel_path.resolve()),
            "source_excel_url": f"{PUBLIC_BASE_URL}/uploads/{user_id}/{excel_stem}/{fname}",
            "row_count": len(items),
            "items": items,
        }
        json_path = user_dir / f"{excel_stem}.json"
        with open(json_path, "w", encoding="utf-8") as jf:
            json.dump(manifest, jf, indent=2)

        result_url = f"{PUBLIC_BASE_URL}/uploads/{user_id}/{excel_stem}.json"
        return JSONResponse({
            "status": "completed",
            "filename": fname,
            "row_count": len(items),
            "result_url": result_url,
            "items": items,
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# -------------------------------- Export → json/csv/xlsx/epub/mobi --------------------------------
@router.post("/export-images")
async def export_excel_alt_text(
    body: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
):
    """
    Exports the Excel/CSV-driven alt-text manifest into json/csv/xlsx/epub/mobi.
    Body must include: user_id, excel_stem, format.
    Optionally: project_id OR (project_name + project_description)

    Optional flags:
      - include_file_name: bool (default False)  # omit "File Name" fields/columns when False
      - allow_stem_fallback: bool (default False)
      - allow_download_if_missing: bool (default False)

    XLSX customisations (optional in body):
      - image_width_px: int (default 240)
      - image_height_px: int (default 200)
      - col_widths: dict with keys "image", "short", "long" (and "file" only if include_file_name=True)
         e.g. {"image": 34, "short": 64, "long": 110}
    """
    try:
        import tempfile, zipfile, csv, uuid, mimetypes, math, io

        PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "http://localhost:8000")

        if not body:
            raise HTTPException(status_code=400, detail="Missing request body")
        if "user_id" not in body or "excel_stem" not in body or "format" not in body:
            raise HTTPException(status_code=400, detail="Missing user_id, excel_stem or format")

        user_id = int(body["user_id"])
        excel_stem = str(body["excel_stem"]).strip()
        fmt = str(body["format"]).lower().strip()
        include_file_name = bool(body.get("include_file_name", False))

        manifest_path = UPLOAD_ROOT / str(user_id) / f"{excel_stem}.json"
        if not manifest_path.exists():
            raise HTTPException(status_code=404, detail="Manifest file not found")

        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f) or {}

        raw_items: List[Dict[str, Any]] = manifest.get("items", []) or []
        items: List[Dict[str, Any]] = []
        for it in raw_items:
            visible = it.get("is_visible", True)
            idx = it.get("index") or it.get("row") or 0
            try:
                idx = int(idx)
            except Exception:
                idx = 0
            if visible and idx > 0:
                items.append(it)

        if not items:
            raise HTTPException(status_code=400, detail="No eligible items found (visible & index > 0)")

        project_id = body.get("project_id")
        if not project_id and body.get("project_name") and body.get("project_description"):
            now = datetime.datetime.utcnow()
            proj = Project(
                owner_id=user_id,
                name=str(body["project_name"]).strip(),
                description=str(body["project_description"]).strip(),
                is_archived=False,
                created_at=now,
                updated_at=now,
            )
            db.add(proj)
            db.commit()
            db.refresh(proj)
            project_id = proj.id

        if project_id:
            proj_row = (
                db.query(Project)
                .filter(Project.id == project_id, Project.owner_id == user_id, Project.is_archived == False)
                .first()
            )
            if not proj_row:
                raise HTTPException(status_code=404, detail="Project not found")

        tmp_path: Optional[str] = None
        download_filename: str = ""
        media_type: str = ""

        def _page(it):  return it.get("page")
        def _idx(it):
            v = it.get("index") or it.get("row")
            try: return int(v)
            except: return v
        def _fname(it):  return (it.get("file_name") or "").strip()
        def _imgurl(it): return (it.get("image_url") or it.get("url") or "").strip()
        def _short(it):  return it.get("short_alt") or it.get("short_alt_text") or ""
        def _long(it):   return it.get("long_alt")  or it.get("long_alt_text")  or ""

        images_dir = UPLOAD_ROOT / str(user_id) / excel_stem / "images"
        ensure_dir(images_dir)

        # ---------- Robust resolver for local image file ----------
        def _resolve_local_image(it) -> Optional[FilePath]:
            """
            Priority:
              1) image_paths[0] if exists
              2) image_url mapped back to local if it's under /uploads/{user}/{excel_stem}/images/
              3) exact file_name (only if include_file_name is True AND present)
              4) stem fallback (ONLY if allow_stem_fallback = true and file_name present)
            """
            # 1) image_paths
            for p in (it.get("image_paths") or []):
                if p and FilePath(p).exists():
                    return FilePath(p)

            # 2) map image_url -> local path (only if URL is inside our uploads dir)
            url = _imgurl(it)
            if url:
                expected_prefix = f"{PUBLIC_BASE_URL.rstrip('/')}/uploads/{user_id}/{excel_stem}/images/"
                if url.startswith(expected_prefix):
                    fn = url[len(expected_prefix):].split("?")[0].split("#")[0]
                    local = images_dir / fn
                    if local.exists():
                        return local

            # 3) exact file_name (only if we keep file names)
            if include_file_name:
                fname = _fname(it)
                if fname:
                    exact = images_dir / fname
                    if exact.exists():
                        return exact

                # 4) optional stem fallback
                if body.get("allow_stem_fallback") and fname:
                    stem = FilePath(fname).stem
                    if stem:
                        for p in images_dir.iterdir():
                            if p.is_file() and p.stem == stem:
                                return p
            return None

        # Optional: last-resort download to local
        def _download_to_local(url: str, stem_hint: str) -> Optional[FilePath]:
            try:
                if not url.lower().startswith(("http://", "https://")):
                    return None
                r = requests.get(url, timeout=20)
                r.raise_for_status()
                data = r.content
                # normalise via PIL
                try:
                    from PIL import Image as PILImage
                    im = PILImage.open(io.BytesIO(data))
                    fmt = (im.format or "PNG").upper()
                    if fmt in ("JPEG", "JPG"):
                        ext = ".jpg"
                        buf = io.BytesIO()
                        im.convert("RGB").save(buf, format="JPEG", quality=92)
                    else:
                        ext = ".png"
                        buf = io.BytesIO()
                        im.save(buf, format="PNG")
                    data = buf.getvalue()
                except Exception:
                    ext = ".png"
                out = images_dir / f"{stem_hint}{ext}"
                with open(out, "wb") as wf:
                    wf.write(data)
                return out
            except Exception:
                return None

        if fmt == "json":
            export_obj = {
                "excel_stem": excel_stem,
                "user_id": user_id,
                "exported_at": datetime.datetime.utcnow().isoformat() + "Z",
                "items": [],
            }
            for it in items:
                row = {
                    "page": _page(it),
                    "index": _idx(it),
                    "image_url": _imgurl(it),
                    "short_alt": _short(it),
                    "long_alt": _long(it),
                }
                if include_file_name:
                    row["file_name"] = _fname(it)
                export_obj["items"].append(row)

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".json")
            with open(tmpf.name, "w", encoding="utf-8") as wf:
                json.dump(export_obj, wf, ensure_ascii=False, indent=2)
            tmp_path = tmpf.name
            download_filename = f"{excel_stem}.json"
            media_type = "application/json"

        elif fmt == "csv":
            headers = ["Page", "Index", "Image URL", "Short Alt", "Long Alt"]
            if include_file_name:
                headers = ["Page", "Index", "File Name", "Image URL", "Short Alt", "Long Alt"]

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
            with open(tmpf.name, "w", encoding="utf-8", newline="") as wf:
                writer = csv.writer(wf)
                writer.writerow(headers)
                for it in items:
                    row = [_page(it), _idx(it)]
                    if include_file_name:
                        row += [_fname(it)]
                    row += [_imgurl(it), _short(it), _long(it)]
                    writer.writerow(row)

            tmp_path = tmpf.name
            download_filename = f"{excel_stem}.csv"
            media_type = "text/csv"

        elif fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            try:
                from openpyxl.drawing.image import Image as XLImage
            except Exception:
                XLImage = None

            if XLImage is None:
                raise HTTPException(
                    status_code=500,
                    detail="Pillow/openpyxl image support is unavailable. Install pillow and openpyxl with image extras.",
                )

            img_w_px = int(body.get("image_width_px", 240))
            img_h_px = int(body.get("image_height_px", 200))

            # Column widths (characters). No "file" unless include_file_name=True.
            default_col_widths = {"image": 34, "short": 64, "long": 110}
            if include_file_name:
                default_col_widths["file"] = 44
            col_widths = {**default_col_widths, **(body.get("col_widths") or {})}

            wb = Workbook()
            ws = wb.active
            ws.title = "AltText"

            if include_file_name:
                headers = ["Page", "Index", "File Name", "Image", "Short Alt", "Long Alt"]
                widths = [8, 8, col_widths["file"], col_widths["image"], col_widths["short"], col_widths["long"]]
                image_col_letter = "D"
                total_cols = 6
            else:
                headers = ["Page", "Index", "Image", "Short Alt", "Long Alt"]
                widths = [8, 8, col_widths["image"], col_widths["short"], col_widths["long"]]
                image_col_letter = "C"
                total_cols = 5

            ws.append(headers)
            hdr_fill = PatternFill("solid", fgColor="F3F4F6")
            for col_idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = hdr_fill
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}1"

            def px_to_points(px: float) -> float:
                return px * 0.75

            def estimate_text_height_points(short_text: str, long_text: str) -> float:
                per_line_pt = 15.0
                short_cpl = max(1, int(col_widths["short"]))
                long_cpl  = max(1, int(col_widths["long"]))
                short_lines = max(1, math.ceil(len(short_text) / (short_cpl * 0.95)))
                long_lines  = max(1, math.ceil(len(long_text) / (long_cpl  * 0.95)))
                return max(short_lines, long_lines) * per_line_pt + 10

            wrap = Alignment(vertical="top", horizontal="left", wrap_text=True)

            for it in items:
                page, idx, s_alt, l_alt = _page(it), _idx(it), _short(it), _long(it)
                if include_file_name:
                    ws.append([page, idx, _fname(it), "", s_alt, l_alt])
                else:
                    ws.append([page, idx, "", s_alt, l_alt])
                row_idx = ws.max_row

                for col_idx in range(1, total_cols + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = wrap

                # Resolve image
                img_loc: Optional[FilePath] = _resolve_local_image(it)
                if (img_loc is None) and body.get("allow_download_if_missing"):
                    url = _imgurl(it)
                    if url:
                        stem_hint = f"p{page}_i{idx}"
                        img_loc = _download_to_local(url, stem_hint=stem_hint)

                if img_loc and img_loc.exists():
                    try:
                        xlimg = XLImage(str(img_loc))
                        xlimg.width = img_w_px
                        xlimg.height = img_h_px
                        ws.add_image(xlimg, f"{image_col_letter}{row_idx}")  # C or D
                    except Exception:
                        pass

                text_h_pt = estimate_text_height_points(s_alt or "", l_alt or "")
                image_h_pt = px_to_points(img_h_px) + 8
                ws.row_dimensions[row_idx].height = max(text_h_pt, image_h_pt, 40.0)

            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmpf.name)
            tmp_path = tmpf.name
            download_filename = f"{excel_stem}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        elif fmt == "epub":
            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".epub")
            with zipfile.ZipFile(tmpf.name, "w") as zf:
                zf.writestr("mimetype", "application/epub+zip", compress_type=zipfile.ZIP_STORED)
                zf.writestr(
                    "META-INF/container.xml",
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<container version="1.0" xmlns="urn:oasis:names:tc:opendocument:xmlns:container">'
                    '<rootfiles><rootfile full-path="OEBPS/content.opf" media-type="application/oebps-package+xml"/></rootfiles>'
                    "</container>",
                )

                manifest_entries = []
                spine_entries = []
                nav_list_items = []
                book_id = str(uuid.uuid4())
                date_iso = datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

                for i, it in enumerate(items, 1):
                    page = _page(it)
                    idx = _idx(it)
                    s_alt = _short(it)
                    l_alt = _long(it)

                    # Resolve local image for embedding
                    img_loc = _resolve_local_image(it)
                    if (img_loc is None) and body.get("allow_download_if_missing"):
                        url = _imgurl(it)
                        if url:
                            stem_hint = f"p{page}_i{idx}"
                            img_loc = _download_to_local(url, stem_hint=stem_hint)

                    embedded_name = None
                    if img_loc and img_loc.exists():
                        embedded_name = f"Images/{img_loc.name}"
                        with open(img_loc, "rb") as fimg:
                            zf.writestr(f"OEBPS/{embedded_name}", fimg.read(), compress_type=zipfile.ZIP_DEFLATED)
                        ext = img_loc.suffix.lower().lstrip(".")
                        media = f"image/{'jpeg' if ext in ['jpg', 'jpeg'] else 'png'}"
                        manifest_entries.append(f'<item id="img{i}" href="{embedded_name}" media-type="{media}"/>')

                    xhtml_id = f"item{i}"
                    xhtml_path = f"OEBPS/Text/page_{i}.xhtml"
                    img_src = f"{embedded_name}" if embedded_name else _imgurl(it)
                    title_txt = f"Page {page} · Figure {idx}"
                    page_html = (
                        '<?xml version="1.0" encoding="utf-8"?>'
                        '<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" lang="en">'
                        '<head><meta charset="utf-8"/><title>Figure</title>'
                        '<style>body{margin:0;padding:0;text-align:center;font-family:system-ui, sans-serif;}'
                        'figure{margin:1rem;}img{max-width:100%;height:auto;}figcaption{font-size:.9rem;opacity:.85}</style>'
                        '</head><body>'
                        f'<figure><img src="{img_src}" alt="{s_alt or title_txt}"/>'
                        f'<figcaption><strong>{title_txt}</strong><br/>{l_alt or title_txt}</figcaption>'
                        '</figure></body></html>'
                    )
                    zf.writestr(xhtml_path, page_html, compress_type=zipfile.ZIP_DEFLATED)
                    manifest_entries.append(f'<item id="{xhtml_id}" href="Text/page_{i}.xhtml" media-type="application/xhtml+xml"/>')
                    spine_entries.append(f'<itemref idref="{xhtml_id}"/>')
                    nav_list_items.append(f'<li><a href="Text/page_{i}.xhtml">{title_txt}</a></li>')

                nav_html = (
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" '
                    'xmlns:epub="http://www.idpf.org/2007/ops" lang="en">'
                    '<head><meta charset="utf-8"/><title>Contents</title></head>'
                    f'<body><nav epub:type="toc" id="toc"><ol>{"".join(nav_list_items)}</ol></nav></body></html>'
                )
                zf.writestr("OEBPS/nav.xhtml", nav_html, compress_type=zipfile.ZIP_DEFLATED)
                manifest_entries.append('<item id="nav" href="nav.xhtml" media-type="application/xhtml+xml" properties="nav"/>')

                content_opf = (
                    '<?xml version="1.0" encoding="utf-8"?>'
                    '<package version="3.0" xmlns="http://www.idpf.org/2007/opf" unique-identifier="bookid">'
                    f'<metadata xmlns:dc="http://purl.org/dc/elements/1.1/">'
                    f'<dc:identifier id="bookid">{book_id}</dc:identifier>'
                    f'<dc:title>{excel_stem}</dc:title><dc:language>en</dc:language><dc:date>{date_iso}</dc:date>'
                    f'</metadata><manifest>{"".join(manifest_entries)}</manifest><spine>{"".join(spine_entries)}</spine></package>'
                )
                zf.writestr("OEBPS/content.opf", content_opf, compress_type=zipfile.ZIP_DEFLATED)

            tmp_path = tmpf.name
            download_filename = f"{excel_stem}.epub"
            media_type = "application/epub+zip"

        elif fmt == "mobi":
            tmpf = tempfile.NamedTemporaryFile(delete=False, suffix=".mobi.zip")
            with zipfile.ZipFile(tmpf.name, "w", zipfile.ZIP_DEFLATED) as zf:
                index_list = []
                for i, it in enumerate(items, 1):
                    page = _page(it)
                    idx = _idx(it)
                    s_alt = _short(it)
                    l_alt = _long(it)

                    # Resolve local image (embed as data URI if we have it)
                    img_loc = _resolve_local_image(it)
                    if (img_loc is None) and body.get("allow_download_if_missing"):
                        url = _imgurl(it)
                        if url:
                            stem_hint = f"p{page}_i{idx}"
                            img_loc = _download_to_local(url, stem_hint=stem_hint)

                    b64 = None
                    mime = "image/png"
                    if img_loc and img_loc.exists():
                        with open(img_loc, "rb") as fimg:
                            b64 = base64.b64encode(fimg.read()).decode("ascii")
                        mime = "image/jpeg" if img_loc.suffix.lower() in (".jpg", ".jpeg") else "image/png"

                    title_txt = f"Page {page} · Figure {idx}"
                    img_src = f"data:{mime};base64,{b64}" if b64 else _imgurl(it)

                    html = (
                        "<!doctype html><html><head><meta charset=\"utf-8\"><title>Figure</title>"
                        "<style>body{margin:0;padding:0;text-align:center;font-family:system-ui, sans-serif;}figure{margin:1rem;}img{max-width:100%;height:auto;}figcaption{font-size:.9rem;opacity:.85}</style>"
                        "</head><body>"
                        f'<figure><img src="{img_src}" alt="{s_alt or title_txt}"/>'
                        f'<figcaption><strong>{title_txt}</strong><br/>{l_alt or title_txt}</figcaption>'
                        "</figure></body></html>"
                    )
                    html_name = f"page_{i}.html"
                    zf.writestr(html_name, html)
                    index_list.append(f'<li><a href="{html_name}">{title_txt}</a></li>')

                toc = (
                    "<!doctype html><html><head><meta charset=\"utf-8\"><title>Contents</title></head>"
                    "<body><h1>Contents</h1><ol>" + "".join(index_list) + "</ol></body></html>"
                )
                zf.writestr("index.html", toc)

            tmp_path = tmpf.name
            download_filename = f"{excel_stem}.mobi.zip"
            media_type = "application/zip"

        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {fmt}")

        export_dir = UPLOAD_ROOT / str(user_id) / excel_stem
        ensure_dir(export_dir)

        final_path = export_dir / download_filename
        try:
            shutil.move(tmp_path, final_path)
        except Exception:
            shutil.copyfile(tmp_path, final_path)
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        public_url = f"{PUBLIC_BASE_URL}/uploads/{user_id}/{excel_stem}/{download_filename}"
        total_images_processed = len(items)
        print("TOTAL IMAGES:", total_images_processed)
        if project_id and ProjectFile is not None:
            try:
                size = final_path.stat().st_size if final_path.exists() else 0
                ext = FilePath(download_filename).suffix.lower().lstrip(".")
                mime_guess, _ = mimetypes.guess_type(download_filename)
                now = datetime.datetime.utcnow()
                pf = ProjectFile(
                    project_id=project_id,
                    uploaded_by=user_id,
                    original_name=download_filename,
                    storage_path=public_url,
                    mime_type=mime_guess or media_type or "application/octet-stream",
                    ext=ext,
                    size_bytes=size,
                    checksum=None,
                    is_deleted=False,
                    created_at=now,
                    updated_at=now,
                )
                db.add(pf)

            except Exception as e:
                 print("PROJECT FILE ERROR:", str(e))

        try:
            logger.info("Fetching user plan")
            user_plan = db.execute(text("""
    SELECT id, image_used
    FROM user_plans
    WHERE user_id = :user_id AND is_active = true
    LIMIT 1
"""), {"user_id": user_id}).fetchone()
            logger.info(f"user_plan: {user_plan}")
            if user_plan:
                logger.info(f"before update image_used: {user_plan.image_used}")
                db.execute(text("""
    UPDATE user_plans
    SET image_used = :image_used,
        updated_at = :updated_at
    WHERE id = :id
"""), {
    "image_used": (user_plan.image_used or 0) + total_images_processed,
    "updated_at": datetime.datetime.utcnow(),
    "id": user_plan.id
})

            db.execute(text("""
    INSERT INTO usage_logs
    (user_id, user_plan_id, type, file_name, credits_used, reference_id, created_at)
    VALUES (:user_id, :user_plan_id, :type, :file_name, :credits_used, :reference_id, :created_at)
"""), {
    "user_id": user_id,
    "user_plan_id": user_plan.id,
    "type": "Ecel to Alt Text Export",
    "file_name": download_filename,
    "credits_used": total_images_processed,
    "reference_id": project_id,
    "created_at": datetime.datetime.utcnow()
})
            logger.info("committing to db")
            db.commit()

        except Exception as e:
            db.rollback()
            logger.error(f"usage log error: {str(e)}", exc_info=True)

        return FileResponse(path=str(final_path), media_type=media_type, filename=download_filename)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
